import pandas as pd
import glob
from openpyxl import load_workbook
import os


class DataFile:

    def __init__(self, file_path):
        self.file_path = file_path

    @staticmethod
    def read_csv_file(file_path):
        """
        reads a *.csv file and returns a pandas DataFrame of the file

        :param file_path: path to a *.csv file
        :return: a pandas DataFrame from the chosen file
        """
        with open(file_path, encoding='UTF-16LE') as file:
            return pd.read_csv(file, sep='\t')

    def combine_to_excel(self, input_directory: str, output_file: str) -> None:
        """
        combines data from csv into one .xlsx file
        :param input_directory: name of directory that contains *.csv files
        :param output_file: name of the new file with combined data
        :return: saves a new *.xlsx file for future manipulations
        """
        parsed = [self.read_csv_file(file_path=path) for path in glob.glob(f'{input_directory}/*.csv')]
        merged = pd.concat(parsed)
        merged.to_excel(output_file, index=False)

    @staticmethod
    def load_invoices(file_path):
        """
        Loads first column in first sheet into a list.
        """
        invoice_file = load_workbook(file_path)
        invoice_sheet = invoice_file.active
        return [
            str(row[0]) for row in
            invoice_sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True)
        ]

    @staticmethod
    def remove_row(file_path):
        """
        Checks if first row contains unnecessary data (not headings)
        and removes it
        :param file_path: path to GFIS excel file

        """
        gfis_file = load_workbook(file_path)
        gfis_sheet = gfis_file.active
        for cell in gfis_sheet.iter_rows(max_row=1, values_only=True):
            if None in cell:
                gfis_sheet.delete_rows(1)
                gfis_file.save(file_path)
                print(f'row has been removed in {file_path}')
            else:
                print(f'no rows to remove in {file_path}. GFIS data spreadsheet is OK.')

    @staticmethod
    def remove_temporary_files(file):
        """
        removes created files after processing
        :param file:
        :return:
        """
        try:
            os.remove(file)
        except FileNotFoundError:
            print('Nothing to remove.')
